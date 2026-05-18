import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from io import BytesIO
from copy import copy


st.set_page_config(
    page_title="나이스 출결 메일머지 엑셀 변환기",
    page_icon="📊",
    layout="wide"
)


# ------------------------------------------------------------
# 기본 설정
# ------------------------------------------------------------

GRADES = [1, 2, 3]

ITEMS_PER_GRADE = [
    "수업일수",
    "질병결석",
    "미인정결석",
    "기타결석",
    "질병지각",
    "미인정지각",
    "기타지각",
    "질병조퇴",
    "미인정조퇴",
    "기타조퇴",
    "질병결과",
    "미인정결과",
    "기타결과",
    "특기사항",
]


def make_output_columns():
    """메일머지용 최종 열 이름을 생성합니다."""
    columns = ["번호", "이름"]

    for grade in GRADES:
        for item in ITEMS_PER_GRADE:
            columns.append(f"{item}_{grade}학년")

    return columns


OUTPUT_COLUMNS = make_output_columns()


def normalize_text(value):
    """셀 값을 비교하기 쉽게 정리합니다."""
    if value is None:
        return ""

    text = str(value).strip()
    text = text.replace(" ", "")
    text = text.replace("\n", "")
    text = text.replace("\t", "")

    return text


def to_number(value, default=0):
    """출결 횟수 값을 숫자로 변환합니다."""
    if value is None:
        return default

    if isinstance(value, (int, float)):
        return int(value)

    text = str(value).strip()

    if text == "":
        return default

    if text in ["-", "·", "."]:
        return default

    try:
        return int(float(text))
    except ValueError:
        return default


def to_text(value):
    """특기사항 값을 문자열로 변환합니다."""
    if value is None:
        return ""

    text = str(value).strip()

    if text.lower() == "none":
        return ""

    return text


# ------------------------------------------------------------
# 병합 셀 처리
# ------------------------------------------------------------

def build_merged_cell_map(ws):
    """
    병합 셀의 하위 셀에서도 왼쪽 위 셀 값을 읽을 수 있도록 매핑을 만듭니다.
    예: B10:B12가 병합되어 있고 B10에 번호가 있으면 B11, B12에서도 같은 번호로 읽음.
    """
    merged_map = {}

    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        top_left_value = ws.cell(row=min_row, column=min_col).value

        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                merged_map[(row, col)] = top_left_value

    return merged_map


def get_cell_value(ws, merged_map, row, col):
    """병합 셀을 고려하여 셀 값을 읽습니다."""
    if (row, col) in merged_map:
        return merged_map[(row, col)]

    return ws.cell(row=row, column=col).value


# ------------------------------------------------------------
# 헤더 자동 감지
# ------------------------------------------------------------

def find_header_rows(ws, merged_map):
    """
    '번 호', '성 명', '학 년'이 있는 행을 헤더 행으로 감지합니다.
    나이스 출력물은 페이지 또는 학생 블록마다 헤더가 반복될 수 있습니다.
    """
    header_rows = []

    for row in range(1, ws.max_row + 1):
        row_values = [
            normalize_text(get_cell_value(ws, merged_map, row, col))
            for col in range(1, ws.max_column + 1)
        ]

        has_number = any(value in ["번호", "번호"] for value in row_values)
        has_name = any(value in ["성명", "성명"] for value in row_values)
        has_grade = any(value in ["학년"] for value in row_values)

        if has_number and has_name and has_grade:
            header_rows.append(row)

    return header_rows


def detect_columns_from_header(ws, merged_map, header_row):
    """
    헤더 행과 그 아래 행을 기준으로 필요한 열 위치를 찾습니다.
    선생님께서 올려주신 나이스 양식처럼
    1행: 결석 / 지각 / 조퇴 / 결과
    2행: 질병 / 미인정 / 기타
    구조를 자동으로 읽습니다.
    """
    subheader_row = header_row + 1

    col_map = {}

    for col in range(1, ws.max_column + 1):
        main = normalize_text(get_cell_value(ws, merged_map, header_row, col))
        sub = normalize_text(get_cell_value(ws, merged_map, subheader_row, col))

        if main == "번호":
            col_map["번호"] = col

        elif main == "성명":
            col_map["이름"] = col

        elif main == "학년":
            col_map["학년"] = col

        elif main == "수업일수":
            col_map["수업일수"] = col

        elif "특기사항" in main:
            col_map["특기사항"] = col

        # 결석, 지각, 조퇴, 결과의 하위 항목 처리
        if main in ["결석", "지각", "조퇴", "결과"] and sub in ["질병", "미인정", "기타"]:
            key = f"{sub}{main}"
            col_map[key] = col

    return col_map


def choose_best_column_map(ws, merged_map, header_rows):
    """
    여러 개의 반복 헤더 중 가장 많은 열을 감지한 헤더를 기준으로 사용합니다.
    """
    best_map = {}
    best_header_row = None

    for header_row in header_rows:
        col_map = detect_columns_from_header(ws, merged_map, header_row)

        if len(col_map) > len(best_map):
            best_map = col_map
            best_header_row = header_row

    return best_header_row, best_map


# ------------------------------------------------------------
# 출결 자료 추출
# ------------------------------------------------------------

def make_empty_student_row(number, name):
    """학생 1명의 기본 행을 만듭니다. 없는 학년 자료는 기본값으로 채웁니다."""
    row = {col: 0 for col in OUTPUT_COLUMNS}
    row["번호"] = number
    row["이름"] = name

    for grade in GRADES:
        row[f"특기사항_{grade}학년"] = ""

    return row


def is_valid_grade(value):
    """학년 값이 1, 2, 3 중 하나인지 확인합니다."""
    if isinstance(value, (int, float)):
        return int(value) in GRADES

    text = normalize_text(value)

    return text in ["1", "2", "3", "1학년", "2학년", "3학년"]


def parse_grade(value):
    """학년 값을 숫자 1, 2, 3으로 변환합니다."""
    if isinstance(value, (int, float)):
        return int(value)

    text = normalize_text(value)
    text = text.replace("학년", "")

    try:
        return int(text)
    except ValueError:
        return None


def extract_attendance_data(ws):
    """
    나이스 출결 엑셀에서 학생별 출결 자료를 추출합니다.
    학생 수, 학년 수가 달라도 자동으로 처리합니다.
    """
    merged_map = build_merged_cell_map(ws)

    header_rows = find_header_rows(ws, merged_map)

    if not header_rows:
        raise ValueError(
            "출결표 헤더를 찾지 못했습니다. "
            "'번 호', '성 명', '학 년'이 있는 나이스 출결상황 양식인지 확인해 주세요."
        )

    header_row, col_map = choose_best_column_map(ws, merged_map, header_rows)

    required_keys = ["번호", "이름", "학년", "수업일수"]
    missing_required = [key for key in required_keys if key not in col_map]

    if missing_required:
        raise ValueError(
            "필수 열을 찾지 못했습니다: "
            + ", ".join(missing_required)
            + "\n원본 엑셀의 헤더 구조를 확인해 주세요."
        )

    students = {}
    detected_rows = 0

    for row_num in range(1, ws.max_row + 1):
        grade_value = get_cell_value(ws, merged_map, row_num, col_map["학년"])

        if not is_valid_grade(grade_value):
            continue

        grade = parse_grade(grade_value)

        if grade not in GRADES:
            continue

        number_value = get_cell_value(ws, merged_map, row_num, col_map["번호"])
        name_value = get_cell_value(ws, merged_map, row_num, col_map["이름"])

        number = to_number(number_value, default=None)
        name = to_text(name_value)

        # 번호와 이름이 없는 행은 학생 자료로 보지 않음
        if number is None or name == "":
            continue

        key = (number, name)

        if key not in students:
            students[key] = make_empty_student_row(number, name)

        student_row = students[key]
        suffix = f"_{grade}학년"

        # 수업일수
        if "수업일수" in col_map:
            student_row[f"수업일수{suffix}"] = to_number(
                get_cell_value(ws, merged_map, row_num, col_map["수업일수"])
            )

        # 결석, 지각, 조퇴, 결과 항목
        count_items = [
            "질병결석", "미인정결석", "기타결석",
            "질병지각", "미인정지각", "기타지각",
            "질병조퇴", "미인정조퇴", "기타조퇴",
            "질병결과", "미인정결과", "기타결과",
        ]

        for item in count_items:
            if item in col_map:
                student_row[f"{item}{suffix}"] = to_number(
                    get_cell_value(ws, merged_map, row_num, col_map[item])
                )

        # 특기사항
        if "특기사항" in col_map:
            student_row[f"특기사항{suffix}"] = to_text(
                get_cell_value(ws, merged_map, row_num, col_map["특기사항"])
            )

        detected_rows += 1

    result_df = pd.DataFrame(list(students.values()), columns=OUTPUT_COLUMNS)

    if not result_df.empty:
        result_df = result_df.sort_values(by=["번호", "이름"]).reset_index(drop=True)

    debug_info = {
        "header_rows": header_rows,
        "selected_header_row": header_row,
        "detected_columns": col_map,
        "detected_student_count": len(students),
        "detected_grade_row_count": detected_rows,
    }

    return result_df, debug_info


# ------------------------------------------------------------
# 엑셀 다운로드 파일 생성
# ------------------------------------------------------------

def dataframe_to_excel_bytes(df):
    """메일머지용 결과 데이터프레임을 엑셀 bytes로 변환합니다."""
    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="메일머지용_출결")

        wb = writer.book
        ws = writer.sheets["메일머지용_출결"]

        # 틀 고정
        ws.freeze_panes = "C2"

        # 필터 적용
        ws.auto_filter.ref = ws.dimensions

        # 열 너비 조정
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter

            for cell in col:
                value = cell.value
                if value is not None:
                    max_length = max(max_length, len(str(value)))

            adjusted_width = min(max(max_length + 2, 8), 24)
            ws.column_dimensions[col_letter].width = adjusted_width

        # 헤더 스타일
        for cell in ws[1]:
            cell.font = copy(cell.font)
            cell.font = cell.font.copy(bold=True)
            cell.alignment = copy(cell.alignment)
            cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

        # 번호 열 가운데 정렬
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            row[0].alignment = row[0].alignment.copy(horizontal="center")

    output.seek(0)
    return output.getvalue()


# ------------------------------------------------------------
# Streamlit 화면
# ------------------------------------------------------------

st.title("📊 나이스 출결 메일머지 엑셀 변환기")

st.write(
    """
    나이스에서 내려받은 **학교생활기록부 출결상황 엑셀 파일**을 업로드하면  
    학생별 1·2·3학년 출결 자료를 한글 메일머지용 엑셀로 정리합니다.
    """
)

st.info(
    """
    결과 엑셀은 항상 1학년, 2학년, 3학년 필드를 모두 포함합니다.  
    원본 파일에 없는 학년 자료는 출결 횟수는 0, 특기사항은 빈칸으로 처리됩니다.
    """
)

uploaded_file = st.file_uploader(
    "나이스 출결상황 엑셀 파일을 업로드하세요.",
    type=["xlsx"]
)

if uploaded_file is not None:
    try:
        workbook_bytes = BytesIO(uploaded_file.read())
        wb = load_workbook(workbook_bytes, data_only=True)

        sheet_names = wb.sheetnames

        selected_sheet = st.selectbox(
            "처리할 시트를 선택하세요.",
            sheet_names,
            index=0
        )

        ws = wb[selected_sheet]

        st.subheader("원본 파일 정보")

        col1, col2, col3 = st.columns(3)

        with col1:
            st.metric("시트명", selected_sheet)

        with col2:
            st.metric("행 수", ws.max_row)

        with col3:
            st.metric("열 수", ws.max_column)

        if st.button("메일머지용 엑셀 만들기", type="primary"):
            with st.spinner("출결 자료를 분석하는 중입니다..."):
                result_df, debug_info = extract_attendance_data(ws)

            if result_df.empty:
                st.error(
                    "학생 출결 자료를 추출하지 못했습니다. "
                    "원본 엑셀 양식이 나이스 학교생활기록부 출결상황 양식인지 확인해 주세요."
                )
                st.stop()

            st.success(
                f"학생 {debug_info['detected_student_count']}명의 자료를 추출했습니다."
            )

            st.subheader("결과 미리보기")
            st.dataframe(result_df, use_container_width=True)

            excel_bytes = dataframe_to_excel_bytes(result_df)

            st.download_button(
                label="메일머지용 출결 엑셀 다운로드",
                data=excel_bytes,
                file_name="메일머지용_출결자료.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

            with st.expander("감지된 열 정보 확인"):
                st.write("감지된 헤더 행:", debug_info["header_rows"])
                st.write("사용한 헤더 행:", debug_info["selected_header_row"])
                st.write("감지된 출결 자료 행 수:", debug_info["detected_grade_row_count"])

                detected_columns_df = pd.DataFrame(
                    [
                        {"항목": key, "열 번호": value}
                        for key, value in debug_info["detected_columns"].items()
                    ]
                )

                st.dataframe(detected_columns_df, use_container_width=True)

            with st.expander("생성되는 메일머지 필드 목록"):
                field_df = pd.DataFrame({"필드명": OUTPUT_COLUMNS})
                st.dataframe(field_df, use_container_width=True)

    except Exception as e:
        st.error("파일을 처리하는 중 오류가 발생했습니다.")
        st.exception(e)

else:
    st.warning("먼저 나이스 출결상황 엑셀 파일을 업로드해 주세요.")
